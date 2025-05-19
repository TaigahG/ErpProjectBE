from decimal import Decimal
from sqlalchemy.orm import Session
from sqlalchemy import and_, func, extract
from datetime import datetime, timedelta
from typing import List, Optional
from models.financial import Transaction, TransactionType
from schemas.financial import AccountCategory
from models.invoice import Invoice, InvoiceStatus
from schemas.reports import ProfitLossReport, BalanceSheet, RevenuePrediction
import pandas as pd
from sklearn.linear_model import LinearRegression
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import datetime




def generate_pnl(db: Session, start_date: datetime, end_date: datetime) -> ProfitLossReport:
    income = db.query(func.sum(Transaction.amount)).filter(
        Transaction.transaction_type == TransactionType.INCOME,
        Transaction.transaction_date >= start_date,
        Transaction.transaction_date <= end_date
    ).scalar() or 0

    expenses = db.query(func.sum(Transaction.amount)).filter(
        Transaction.transaction_type == TransactionType.EXPENSE,
        Transaction.transaction_date >= start_date,
        Transaction.transaction_date <= end_date
    ).scalar() or 0

    revenue_breakdown = db.query(
        func.date_trunc('month', Transaction.transaction_date).label('month'),
        func.sum(Transaction.amount).label('amount')
    ).filter(
        Transaction.transaction_type == TransactionType.INCOME,
        Transaction.transaction_date >= start_date,
        Transaction.transaction_date <= end_date
    ).group_by(
        func.date_trunc('month', Transaction.transaction_date)
    ).order_by(
        func.date_trunc('month', Transaction.transaction_date)
    ).all()

    expense_breakdown = db.query(
        func.date_trunc('month', Transaction.transaction_date).label('month'),
        func.sum(Transaction.amount).label('amount')
    ).filter(
        Transaction.transaction_type == TransactionType.EXPENSE,
        Transaction.transaction_date >= start_date,
        Transaction.transaction_date <= end_date
    ).group_by(
        func.date_trunc('month', Transaction.transaction_date)
    ).order_by(
        func.date_trunc('month', Transaction.transaction_date)
    ).all()

    total_revenue = sum(r.amount for r in revenue_breakdown)
    total_expenses = sum(e.amount for e in expense_breakdown)

    return ProfitLossReport(
        period_start=start_date,
        period_end=end_date,
        total_revenue=total_revenue,
        total_expenses=total_expenses,
        net_profit=total_revenue - total_expenses,
        revenue_breakdown=[{
            "category": r.month.strftime("%Y-%m-%d"),
            "amount": float(r.amount)
        } for r in revenue_breakdown],
        expenses_breakdown=[{
            "category": e.month.strftime("%Y-%m-%d"),
            "amount": float(e.amount)
        } for e in expense_breakdown]
    )

def generate_balance_sheet(db: Session, as_of_date: datetime) -> BalanceSheet:
    total_assets = db.query(func.sum(Transaction.amount)).filter(
        Transaction.transaction_type == TransactionType.INCOME,
        Transaction.transaction_type == TransactionType.ASSET,
        Transaction.transaction_date <= as_of_date
    ).scalar() or 0

    total_assets_1 = db.query(func.sum(Transaction.amount)).filter(
        Transaction.transaction_type == TransactionType.ASSET,
        Transaction.transaction_date <= as_of_date
    ).scalar() or 0

    total_liabilities = db.query(func.sum(Transaction.amount)).filter(
        Transaction.transaction_type == TransactionType.EXPENSE,
        Transaction.transaction_date <= as_of_date
    ).scalar() or 0

    total_liabilities_1 = db.query(func.sum(Transaction.amount)).filter(
        Transaction.transaction_type == TransactionType.LIABILITY,
        Transaction.transaction_date <= as_of_date
    ).scalar() or 0
    

    return BalanceSheet(
        as_of_date=as_of_date,
        total_assets=total_assets_1,
        total_liabilities=total_liabilities_1,
        total_equity=total_assets - total_liabilities,
        assets_breakdown=[], 
        liabilities_breakdown=[],
        equity_breakdown=[]
    )

def predict_revenue(db: Session, months_ahead: int = 3) -> List[RevenuePrediction]:
    monthly_revenue = db.query(
        func.date_trunc('month', Transaction.transaction_date).label('month'),
        func.sum(Transaction.amount).label('revenue')
    ).filter(
        Transaction.transaction_type == TransactionType.INCOME
    ).group_by(
        func.date_trunc('month', Transaction.transaction_date)
    ).order_by('month').all()

    if not monthly_revenue:
        return []

    df = pd.DataFrame(monthly_revenue, columns=['month', 'revenue'])
    df['revenue'] = df['revenue'].astype(float)
    
    last_month = df['month'].max()

    X = np.arange(len(df)).reshape(-1, 1)
    y = df['revenue'].values  

    model = LinearRegression()
    model.fit(X, y)

    y_pred = model.predict(X)
    
    mse = np.mean((y - y_pred) ** 2)
    
    r2 = model.score(X, y)
    print(f"Model R²: {r2:.2f}")
    
    n = len(X)
    std_error = np.sqrt(mse / (n - 2)) if n > 2 else 0
    
    predictions = []
    for i in range(1, months_ahead + 1):
        next_month = (last_month + pd.DateOffset(months=i)).replace(day=1)
        
        month_index = len(df) + i - 1
        X_new = np.array([[month_index]])
        predicted_value = float(model.predict(X_new)[0])
        
        time_factor = 1 / (i + 1) 
        data_quality = min(1.0, len(df) / 12)  
        model_accuracy = max(0, r2)  

        
        confidence_level = (time_factor * 0.3 + 
                          data_quality * 0.3 + 
                          model_accuracy * 0.4)
        
        prediction_dict = {
            "prediction_date": next_month,
            "predicted_amount": max(Decimal(str(predicted_value)), Decimal('0')),
            "confidence_level": round(confidence_level, 2),
            "factors": [
                f"Model accuracy (R²): {r2:.2f}",
                f"Data points: {len(df)} months",
                f"Prediction distance: {i} months",
                f"Standard error: {std_error:.2f}"
            ]
        }
        predictions.append(RevenuePrediction(**prediction_dict))

    return predictions

def get_dashboard_data(db: Session, period: str) -> dict:
    latest_transaction = db.query(func.max(Transaction.transaction_date)).scalar()
    if not latest_transaction:
        return {
            "total_income": 0,
            "total_expenses": 0,
            "net_profit": 0,
            "previous_income": 0,
            "previous_expenses": 0,
            "previous_profit": 0,
            "monthly_data": []
        }

    reference_date = latest_transaction

    if period == '30d':
        end_date = reference_date
        start_date = reference_date.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        previous_start = (start_date - timedelta(days=1)).replace(day=1)
    elif period == '90d':
        end_date = reference_date
        start_date = (reference_date - timedelta(days=90)).replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        previous_start = (start_date - timedelta(days=90)).replace(day=1)
    else:  
        end_date = reference_date
        start_date = datetime(reference_date.year, 1, 1, 0, 0, 0)
        previous_start = datetime(reference_date.year - 1, 1, 1, 0, 0, 0)

    print(f"Period selected: {period}")
    print(f"Reference date: {reference_date}")
    print(f"Start date: {start_date}")
    print(f"End date: {end_date}")

    current_data = get_period_data(db, start_date, end_date)
    previous_data = get_period_data(db, previous_start, start_date)
    monthly_data = get_monthly_breakdown(db, start_date, end_date)

    return {
        "total_income": current_data["income"],
        "total_expenses": current_data["expenses"],
        "net_profit": current_data["income"] - current_data["expenses"],
        "previous_income": previous_data["income"],
        "previous_expenses": previous_data["expenses"],
        "previous_profit": previous_data["income"] - previous_data["expenses"],
        "monthly_data": monthly_data
    }

def get_period_data(db: Session, start_date: datetime, end_date: datetime) -> dict:
    print(f"Querying transactions from {start_date} to {end_date}")

    transactions = db.query(Transaction).filter(
        Transaction.transaction_date >= start_date,
        Transaction.transaction_date <= end_date
    ).all()

    income = sum(t.amount for t in transactions if t.transaction_type == TransactionType.INCOME)
    expenses = sum(t.amount for t in transactions if t.transaction_type == TransactionType.EXPENSE)

    print(f"Found {len(transactions)} transactions")
    print(f"Total income: {income}")
    print(f"Total expenses: {expenses}")

    return {
        "income": float(income),
        "expenses": float(expenses)
    }
def get_monthly_breakdown(db: Session, start_date: datetime, end_date: datetime) -> List[dict]:

    start_date = start_date.replace(hour=0, minute=0, second=0, microsecond=0)
    end_date = end_date.replace(hour=23, minute=59, second=59, microsecond=999999)

    income_by_month = db.query(
        func.date_trunc('month', Transaction.transaction_date).label('month'),
        func.sum(Transaction.amount).label('income')
    ).filter(
        Transaction.transaction_type == TransactionType.INCOME,
        Transaction.transaction_date >= start_date,
        Transaction.transaction_date <= end_date
    ).group_by('month').all()

    expenses_by_month = db.query(
        func.date_trunc('month', Transaction.transaction_date).label('month'),
        func.sum(Transaction.amount).label('expenses')
    ).filter(
        Transaction.transaction_type == TransactionType.EXPENSE,
        Transaction.transaction_date >= start_date,
        Transaction.transaction_date <= end_date
    ).group_by('month').all()

    monthly_data = {}
    for income in income_by_month:
        month_key = income.month.strftime("%b %Y")
        monthly_data[month_key] = {
            "month": month_key,
            "income": float(income.income),
            "expenses": 0
        }

    for expense in expenses_by_month:
        month_key = expense.month.strftime("%b %Y")
        if month_key in monthly_data:
            monthly_data[month_key]["expenses"] = float(expense.expenses)
        else:
            monthly_data[month_key] = {
                "month": month_key,
                "income": 0,
                "expenses": float(expense.expenses)
            }

    return list(monthly_data.values())


def generate_balance_sheet_ifrs(db: Session, as_of_date: datetime) -> dict:
    """
    Generate a balance sheet in IFRS format
    """
    # Get asset accounts and their balances
    assets_query = db.query(
        AccountCategory.name.label('category'),
        AccountCategory.code.label('code'),
        AccountCategory.parent_id.label('parent_id'),
        func.sum(Transaction.amount).label('amount')
    ).outerjoin(
        Transaction, 
        and_(
            Transaction.account_category_id == AccountCategory.id,
            Transaction.transaction_date <= as_of_date
        )
    ).filter(
        AccountCategory.type == TransactionType.ASSET
    ).group_by(
        AccountCategory.id,
        AccountCategory.name,
        AccountCategory.code,
        AccountCategory.parent_id
    ).order_by(AccountCategory.code)
    
    assets = [{
        'category': row.category,
        'code': row.code,
        'parent_id': row.parent_id,
        'amount': float(row.amount) if row.amount else 0
    } for row in assets_query.all()]
    
    # Similar queries for liabilities and equity
    liabilities_query = db.query(
        AccountCategory.name.label('category'),
        AccountCategory.code.label('code'),
        AccountCategory.parent_id.label('parent_id'),
        func.sum(Transaction.amount).label('amount')
    ).outerjoin(
        Transaction, 
        and_(
            Transaction.account_category_id == AccountCategory.id,
            Transaction.transaction_date <= as_of_date
        )
    ).filter(
        AccountCategory.type == TransactionType.LIABILITY
    ).group_by(
        AccountCategory.id,
        AccountCategory.name,
        AccountCategory.code,
        AccountCategory.parent_id
    ).order_by(AccountCategory.code)
    
    liabilities = [{
        'category': row.category,
        'code': row.code,
        'parent_id': row.parent_id,
        'amount': float(row.amount) if row.amount else 0
    } for row in liabilities_query.all()]
    
    equity_query = db.query(
        AccountCategory.name.label('category'),
        AccountCategory.code.label('code'),
        AccountCategory.parent_id.label('parent_id'),
        func.sum(Transaction.amount).label('amount')
    ).outerjoin(
        Transaction, 
        and_(
            Transaction.account_category_id == AccountCategory.id,
            Transaction.transaction_date <= as_of_date
        )
    ).filter(
        AccountCategory.type == TransactionType.EQUITY
    ).group_by(
        AccountCategory.id,
        AccountCategory.name,
        AccountCategory.code,
        AccountCategory.parent_id
    ).order_by(AccountCategory.code)
    
    equity = [{
        'category': row.category,
        'code': row.code,
        'parent_id': row.parent_id,
        'amount': float(row.amount) if row.amount else 0
    } for row in equity_query.all()]
    
    # Calculate totals
    total_assets = sum(item['amount'] for item in assets)
    total_liabilities = sum(item['amount'] for item in liabilities)
    total_equity = sum(item['amount'] for item in equity)
    
    # Build hierarchical structure for assets, liabilities, and equity
    assets_tree = build_account_tree(assets)
    liabilities_tree = build_account_tree(liabilities)
    equity_tree = build_account_tree(equity)
    
    return {
        'as_of_date': as_of_date.isoformat(),
        'total_assets': total_assets,
        'total_liabilities': total_liabilities,
        'total_equity': total_equity,
        'assets': assets_tree,
        'liabilities': liabilities_tree,
        'equity': equity_tree
    }

def generate_pnl_ifrs(db: Session, start_date: datetime, end_date: datetime) -> dict:
    """
    Generate a profit and loss statement in IFRS format
    """
    # Get revenue accounts and their balances
    revenue_query = db.query(
        AccountCategory.name.label('category'),
        AccountCategory.code.label('code'),
        AccountCategory.parent_id.label('parent_id'),
        func.sum(Transaction.amount).label('amount')
    ).outerjoin(
        Transaction, 
        and_(
            Transaction.account_category_id == AccountCategory.id,
            Transaction.transaction_date >= start_date,
            Transaction.transaction_date <= end_date
        )
    ).filter(
        AccountCategory.type == TransactionType.REVENUE
    ).group_by(
        AccountCategory.id,
        AccountCategory.name,
        AccountCategory.code,
        AccountCategory.parent_id
    ).order_by(AccountCategory.code)
    
    revenue = [{
        'category': row.category,
        'code': row.code,
        'parent_id': row.parent_id,
        'amount': float(row.amount) if row.amount else 0
    } for row in revenue_query.all()]
    
    # Get expense accounts and their balances
    expense_query = db.query(
        AccountCategory.name.label('category'),
        AccountCategory.code.label('code'),
        AccountCategory.parent_id.label('parent_id'),
        func.sum(Transaction.amount).label('amount')
    ).outerjoin(
        Transaction, 
        and_(
            Transaction.account_category_id == AccountCategory.id,
            Transaction.transaction_date >= start_date,
            Transaction.transaction_date <= end_date
        )
    ).filter(
        AccountCategory.type == TransactionType.EXPENSE
    ).group_by(
        AccountCategory.id,
        AccountCategory.name,
        AccountCategory.code,
        AccountCategory.parent_id
    ).order_by(AccountCategory.code)
    
    expenses = [{
        'category': row.category,
        'code': row.code,
        'parent_id': row.parent_id,
        'amount': float(row.amount) if row.amount else 0
    } for row in expense_query.all()]
    
    total_revenue = sum(item['amount'] for item in revenue)
    total_expenses = sum(item['amount'] for item in expenses)
    net_profit = total_revenue - total_expenses
    
    revenue_tree = build_account_tree(revenue)
    expenses_tree = build_account_tree(expenses)
    
    return {
        'period_start': start_date.isoformat(),
        'period_end': end_date.isoformat(),
        'total_revenue': total_revenue,
        'total_expenses': total_expenses,
        'net_profit': net_profit,
        'revenue': revenue_tree,
        'expenses': expenses_tree
    }

def build_account_tree(accounts):
    """
    Build a hierarchical tree structure from a flat list of accounts
    """
    # Create a map of id to account
    id_map = {}
    for acc in accounts:
        acc_id = acc.get('id')
        if acc_id:
            id_map[acc_id] = acc
            acc['children'] = []
    
    # Organize into a tree
    root_accounts = []
    for acc in accounts:
        parent_id = acc.get('parent_id')
        if parent_id is None:
            root_accounts.append(acc)
        elif parent_id in id_map:
            parent = id_map[parent_id]
            if 'children' not in parent:
                parent['children'] = []
            parent['children'].append(acc)
    
    return root_accounts

def generate_pdf_report(report_type, data):
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=letter,
        rightMargin=72,
        leftMargin=72,
        topMargin=72,
        bottomMargin=72
    )
    elements = []
    styles = getSampleStyleSheet()
    
    # Create custom styles
    title_style = ParagraphStyle(
        'TitleStyle',
        parent=styles['Title'],
        fontSize=16,
        textColor=colors.navy,
        spaceAfter=12
    )
    
    subtitle_style = ParagraphStyle(
        'SubtitleStyle',
        parent=styles['Heading2'],
        fontSize=12,
        textColor=colors.navy,
        spaceAfter=6
    )
    
    heading_style = ParagraphStyle(
        'HeadingStyle',
        parent=styles['Heading3'],
        fontSize=12,
        textColor=colors.navy,
        spaceBefore=12,
        spaceAfter=6
    )
    
    # Add logo and company name
    elements.append(Paragraph("PT. Jurnal by Mekari", title_style))
    
    # Add title and date information
    if report_type == "profit-loss-ifrs":
        title = "Laporan Neraca Laba Rugi"
        period_start = datetime.datetime.fromisoformat(data['period_start'].replace('Z', '+00:00'))
        period_end = datetime.datetime.fromisoformat(data['period_end'].replace('Z', '+00:00'))
        subtitle = f"Periode {period_start.strftime('%d %B %Y')} s/d {period_end.strftime('%d %B %Y')}"
    else:  # balance-sheet-ifrs
        title = "Laporan Neraca Keuangan"
        as_of_date = datetime.datetime.fromisoformat(data['as_of_date'].replace('Z', '+00:00'))
        subtitle = f"Per {as_of_date.strftime('%d %B %Y')}"
    
    elements.append(Paragraph(title, title_style))
    elements.append(Paragraph(subtitle, subtitle_style))
    elements.append(Spacer(1, 20))
    
    # Add report data
    if report_type == "profit-loss-ifrs":
        # Add revenue section
        elements.append(Paragraph("Pendapatan Penjualan", heading_style))
        
        # Create table for revenue
        revenue_data = []
        
        # Format the hierarchical revenue data
        def add_revenue_rows(items, level=0):
            for item in items:
                indent = '    ' * level
                revenue_data.append([
                    f"{indent}{item['code']} - {item['category']}", 
                    f"Rp {item['amount']:,.2f}"
                ])
                if item.get('children'):
                    add_revenue_rows(item['children'], level + 1)
        
        add_revenue_rows(data['revenue'])
        revenue_data.append(["Total Pendapatan", f"Rp {data['total_revenue']:,.2f}"])
        
        # Create and style revenue table
        revenue_table = Table(revenue_data, colWidths=[4*inch, 2*inch])
        revenue_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('LINEABOVE', (0, -1), (-1, -1), 1, colors.black),
            ('LINEBELOW', (0, -1), (-1, -1), 1, colors.black),
        ]))
        elements.append(revenue_table)
        elements.append(Spacer(1, 15))
        
        # Add expenses section
        elements.append(Paragraph("Beban Operasi", heading_style))
        
        # Create table for expenses
        expense_data = []
        
        # Format the hierarchical expense data
        def add_expense_rows(items, level=0):
            for item in items:
                indent = '    ' * level
                expense_data.append([
                    f"{indent}{item['code']} - {item['category']}", 
                    f"Rp {item['amount']:,.2f}"
                ])
                if item.get('children'):
                    add_expense_rows(item['children'], level + 1)
        
        add_expense_rows(data['expenses'])
        expense_data.append(["Total Beban", f"Rp {data['total_expenses']:,.2f}"])
        
        # Create and style expenses table
        expense_table = Table(expense_data, colWidths=[4*inch, 2*inch])
        expense_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('LINEABOVE', (0, -1), (-1, -1), 1, colors.black),
            ('LINEBELOW', (0, -1), (-1, -1), 1, colors.black),
        ]))
        elements.append(expense_table)
        elements.append(Spacer(1, 15))
        
        # Add net profit
        net_profit_data = [["Laba Bersih", f"Rp {data['net_profit']:,.2f}"]]
        net_profit_table = Table(net_profit_data, colWidths=[4*inch, 2*inch])
        net_profit_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('LINEABOVE', (0, 0), (-1, 0), 1, colors.black),
            ('LINEBELOW', (0, 0), (-1, 0), 1, colors.black),
        ]))
        elements.append(net_profit_table)
        
    else:  # balance-sheet-ifrs
        # Add assets section
        elements.append(Paragraph("ASET", heading_style))
        
        asset_data = []
        
        def add_asset_rows(items, level=0):
            for item in items:
                indent = '    ' * level
                asset_data.append([
                    f"{indent}{item['code']} - {item['category']}", 
                    f"Rp {item['amount']:,.2f}"
                ])
                if item.get('children'):
                    add_asset_rows(item['children'], level + 1)
        
        add_asset_rows(data['assets'])
        asset_data.append(["Total Aset", f"Rp {data['total_assets']:,.2f}"])
        
        # Create and style assets table
        asset_table = Table(asset_data, colWidths=[4*inch, 2*inch])
        asset_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('LINEABOVE', (0, -1), (-1, -1), 1, colors.black),
            ('LINEBELOW', (0, -1), (-1, -1), 1, colors.black),
        ]))
        elements.append(asset_table)
        elements.append(Spacer(1, 15))
        
        # Add liabilities section
        elements.append(Paragraph("LIABILITAS", heading_style))
        
        # Create table for liabilities
        liability_data = []
        
        # Format the hierarchical liability data
        def add_liability_rows(items, level=0):
            for item in items:
                indent = '    ' * level
                liability_data.append([
                    f"{indent}{item['code']} - {item['category']}", 
                    f"Rp {item['amount']:,.2f}"
                ])
                if item.get('children'):
                    add_liability_rows(item['children'], level + 1)
        
        add_liability_rows(data['liabilities'])
        liability_data.append(["Total Liabilitas", f"Rp {data['total_liabilities']:,.2f}"])
        
        # Create and style liabilities table
        liability_table = Table(liability_data, colWidths=[4*inch, 2*inch])
        liability_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('LINEABOVE', (0, -1), (-1, -1), 1, colors.black),
            ('LINEBELOW', (0, -1), (-1, -1), 1, colors.black),
        ]))
        elements.append(liability_table)
        elements.append(Spacer(1, 15))
        
        # Add equity section
        elements.append(Paragraph("EKUITAS", heading_style))
        
        # Create table for equity
        equity_data = []
        
        # Format the hierarchical equity data
        def add_equity_rows(items, level=0):
            for item in items:
                indent = '    ' * level
                equity_data.append([
                    f"{indent}{item['code']} - {item['category']}", 
                    f"Rp {item['amount']:,.2f}"
                ])
                if item.get('children'):
                    add_equity_rows(item['children'], level + 1)
        
        add_equity_rows(data['equity'])
        equity_data.append(["Total Ekuitas", f"Rp {data['total_equity']:,.2f}"])
        
        # Create and style equity table
        equity_table = Table(equity_data, colWidths=[4*inch, 2*inch])
        equity_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('LINEABOVE', (0, -1), (-1, -1), 1, colors.black),
            ('LINEBELOW', (0, -1), (-1, -1), 1, colors.black),
        ]))
        elements.append(equity_table)
        elements.append(Spacer(1, 15))
        
        # Add total liabilities and equity
        liab_equity_data = [["Total Liabilitas dan Ekuitas", f"Rp {data['total_liabilities'] + data['total_equity']:,.2f}"]]
        liab_equity_table = Table(liab_equity_data, colWidths=[4*inch, 2*inch])
        liab_equity_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('LINEABOVE', (0, 0), (-1, 0), 1, colors.black),
            ('LINEBELOW', (0, 0), (-1, 0), 1, colors.black),
        ]))
        elements.append(liab_equity_table)
    
    def add_page_number(canvas, doc):
        page_num = canvas.getPageNumber()
        text = f"Page {page_num}"
        canvas.drawRightString(letter[0]-0.5*inch, 0.5*inch, text)
        canvas.drawString(0.5*inch, 0.5*inch, "MaleoAI")
    
    doc.build(elements)
    pdf_data = buffer.getvalue()
    buffer.close()
    return pdf_data

def generate_excel_report(report_type, data):
    wb = openpyxl.Workbook()
    ws = wb.active
    
    title_font = Font(name='Arial', size=14, bold=True, color='000080')  # Navy blue
    subtitle_font = Font(name='Arial', size=12, bold=True, color='000080')
    header_font = Font(name='Arial', size=11, bold=True)
    total_font = Font(name='Arial', size=11, bold=True)
    normal_font = Font(name='Arial', size=10)
    
    header_fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
    total_fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Set title and subtitle
    ws["A1"] = "PT. Jurnal by Mekari"
    ws["A1"].font = title_font
    
    if report_type == "profit-loss-ifrs":
        ws.title = "Laba Rugi"
        ws["A2"] = "Laporan Neraca Laba Rugi"
        period_start = datetime.datetime.fromisoformat(data['period_start'].replace('Z', '+00:00'))
        period_end = datetime.datetime.fromisoformat(data['period_end'].replace('Z', '+00:00'))
        ws["A3"] = f"Periode {period_start.strftime('%d %B %Y')} s/d {period_end.strftime('%d %B %Y')}"
    else:  # balance-sheet-ifrs
        ws.title = "Neraca"
        ws["A2"] = "Laporan Neraca Keuangan"
        as_of_date = datetime.datetime.fromisoformat(data['as_of_date'].replace('Z', '+00:00'))
        ws["A3"] = f"Per {as_of_date.strftime('%d %B %Y')}"
    
    ws["A2"].font = subtitle_font
    ws["A3"].font = subtitle_font
    
    # Start row for data
    current_row = 5
    
    if report_type == "profit-loss-ifrs":
        # Add revenue section
        ws[f"A{current_row}"] = "PENDAPATAN PENJUALAN"
        ws[f"A{current_row}"].font = header_font
        current_row += 1
        
        # Function to add hierarchical items
        def add_items_to_sheet(items, start_row, level=0):
            row = start_row
            
            for item in items:
                indent = '    ' * level
                ws[f"A{row}"] = f"{indent}{item['code']} - {item['category']}"
                ws[f"B{row}"] = item['amount']
                ws[f"B{row}"].number_format = '#,##0.00'
                
                # Apply styles
                ws[f"A{row}"].font = normal_font
                ws[f"B{row}"].font = normal_font
                ws[f"A{row}"].alignment = Alignment(horizontal='left')
                ws[f"B{row}"].alignment = Alignment(horizontal='right')
                
                row += 1
                
                if item.get('children'):
                    row = add_items_to_sheet(item['children'], row, level + 1)
            
            return row
        
        # Add revenue items
        current_row = add_items_to_sheet(data['revenue'], current_row)
        
        # Add total revenue
        ws[f"A{current_row}"] = "Total Pendapatan"
        ws[f"B{current_row}"] = data['total_revenue']
        ws[f"B{current_row}"].number_format = '#,##0.00'
        
        # Apply styles to total
        ws[f"A{current_row}"].font = total_font
        ws[f"B{current_row}"].font = total_font
        ws[f"A{current_row}"].alignment = Alignment(horizontal='left')
        ws[f"B{current_row}"].alignment = Alignment(horizontal='right')
        for col in ['A', 'B']:
            ws[f"{col}{current_row}"].fill = total_fill
            ws[f"{col}{current_row}"].border = border
        
        current_row += 2
        
        # Add expenses section
        ws[f"A{current_row}"] = "BEBAN OPERASI"
        ws[f"A{current_row}"].font = header_font
        current_row += 1
        
        # Add expense items
        current_row = add_items_to_sheet(data['expenses'], current_row)
        
        # Add total expenses
        ws[f"A{current_row}"] = "Total Beban"
        ws[f"B{current_row}"] = data['total_expenses']
        ws[f"B{current_row}"].number_format = '#,##0.00'
        
        # Apply styles to total
        ws[f"A{current_row}"].font = total_font
        ws[f"B{current_row}"].font = total_font
        ws[f"A{current_row}"].alignment = Alignment(horizontal='left')
        ws[f"B{current_row}"].alignment = Alignment(horizontal='right')
        for col in ['A', 'B']:
            ws[f"{col}{current_row}"].fill = total_fill
            ws[f"{col}{current_row}"].border = border
        
        current_row += 2
        
        # Add net profit
        ws[f"A{current_row}"] = "LABA BERSIH"
        ws[f"B{current_row}"] = data['net_profit']
        ws[f"B{current_row}"].number_format = '#,##0.00'
        
        # Apply styles to net profit
        ws[f"A{current_row}"].font = Font(name='Arial', size=12, bold=True)
        ws[f"B{current_row}"].font = Font(name='Arial', size=12, bold=True)
        ws[f"A{current_row}"].alignment = Alignment(horizontal='left')
        ws[f"B{current_row}"].alignment = Alignment(horizontal='right')
        for col in ['A', 'B']:
            ws[f"{col}{current_row}"].fill = PatternFill(start_color='D0D0D0', end_color='D0D0D0', fill_type='solid')
            ws[f"{col}{current_row}"].border = border
    
    else:  # balance-sheet-ifrs
        # Add assets section
        ws[f"A{current_row}"] = "ASET"
        ws[f"A{current_row}"].font = header_font
        current_row += 1
        
        # Function to add hierarchical items
        def add_items_to_sheet(items, start_row, level=0):
            row = start_row
            
            for item in items:
                indent = '    ' * level
                ws[f"A{row}"] = f"{indent}{item['code']} - {item['category']}"
                ws[f"B{row}"] = item['amount']
                ws[f"B{row}"].number_format = '#,##0.00'
                
                # Apply styles
                ws[f"A{row}"].font = normal_font
                ws[f"B{row}"].font = normal_font
                ws[f"A{row}"].alignment = Alignment(horizontal='left')
                ws[f"B{row}"].alignment = Alignment(horizontal='right')
                
                row += 1
                
                if item.get('children'):
                    row = add_items_to_sheet(item['children'], row, level + 1)
            
            return row
        
        # Add asset items
        current_row = add_items_to_sheet(data['assets'], current_row)
        
        # Add total assets
        ws[f"A{current_row}"] = "Total Aset"
        ws[f"B{current_row}"] = data['total_assets']
        ws[f"B{current_row}"].number_format = '#,##0.00'
        
        # Apply styles to total
        ws[f"A{current_row}"].font = total_font
        ws[f"B{current_row}"].font = total_font
        ws[f"A{current_row}"].alignment = Alignment(horizontal='left')
        ws[f"B{current_row}"].alignment = Alignment(horizontal='right')
        for col in ['A', 'B']:
            ws[f"{col}{current_row}"].fill = total_fill
            ws[f"{col}{current_row}"].border = border
        
        current_row += 2
        
        # Add liabilities section
        ws[f"A{current_row}"] = "LIABILITAS"
        ws[f"A{current_row}"].font = header_font
        current_row += 1
        
        # Add liability items
        current_row = add_items_to_sheet(data['liabilities'], current_row)
        
        # Add total liabilities
        ws[f"A{current_row}"] = "Total Liabilitas"
        ws[f"B{current_row}"] = data['total_liabilities']
        ws[f"B{current_row}"].number_format = '#,##0.00'
        
        # Apply styles to total
        ws[f"A{current_row}"].font = total_font
        ws[f"B{current_row}"].font = total_font
        ws[f"A{current_row}"].alignment = Alignment(horizontal='left')
        ws[f"B{current_row}"].alignment = Alignment(horizontal='right')
        for col in ['A', 'B']:
            ws[f"{col}{current_row}"].fill = total_fill
            ws[f"{col}{current_row}"].border = border
        
        current_row += 2
        
        # Add equity section
        ws[f"A{current_row}"] = "EKUITAS"
        ws[f"A{current_row}"].font = header_font
        current_row += 1
        
        # Add equity items
        current_row = add_items_to_sheet(data['equity'], current_row)
        
        # Add total equity
        ws[f"A{current_row}"] = "Total Ekuitas"
        ws[f"B{current_row}"] = data['total_equity']
        ws[f"B{current_row}"].number_format = '#,##0.00'
        
        # Apply styles to total
        ws[f"A{current_row}"].font = total_font
        ws[f"B{current_row}"].font = total_font
        ws[f"A{current_row}"].alignment = Alignment(horizontal='left')
        ws[f"B{current_row}"].alignment = Alignment(horizontal='right')
        for col in ['A', 'B']:
            ws[f"{col}{current_row}"].fill = total_fill
            ws[f"{col}{current_row}"].border = border
        
        current_row += 2
        
        # Add total liabilities and equity
        ws[f"A{current_row}"] = "TOTAL LIABILITAS DAN EKUITAS"
        ws[f"B{current_row}"] = data['total_liabilities'] + data['total_equity']
        ws[f"B{current_row}"].number_format = '#,##0.00'
        
        # Apply styles to final total
        ws[f"A{current_row}"].font = Font(name='Arial', size=12, bold=True)
        ws[f"B{current_row}"].font = Font(name='Arial', size=12, bold=True)
        ws[f"A{current_row}"].alignment = Alignment(horizontal='left')
        ws[f"B{current_row}"].alignment = Alignment(horizontal='right')
        for col in ['A', 'B']:
            ws[f"{col}{current_row}"].fill = PatternFill(start_color='D0D0D0', end_color='D0D0D0', fill_type='solid')
            ws[f"{col}{current_row}"].border = border
    
    # Set column widths
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 20
    
    buffer = BytesIO()
    wb.save(buffer)
    excel_data = buffer.getvalue()
    buffer.close()
    return excel_data

def generate_anthropic_report(report_type, data):
    """
    Use Anthropic API to generate an enhanced report design
    """
    import anthropic
    import json
    from io import BytesIO
    import base64
    import os
    
    api_key = os.getenv("ANTHROPIC_API_KEY")
    client = anthropic.Anthropic(api_key=api_key)
    
    # Convert data to JSON string for the prompt
    data_json = json.dumps(data, indent=2)
    
    # Create the prompt based on report type
    if report_type == "profit-loss-ifrs":
        prompt = f"""
        Create a visually appealing Profit & Loss Statement in an HTML format that I can save as a PDF. 
        The report should look professional and use a clean design with the following data:
        
        {data_json}
        
        The report should include:
        1. Company header with "PT. Jurnal by Mekari" as the company name
        2. Title "Laporan Neraca Laba Rugi"
        3. Period information showing start and end dates
        4. Revenue section with hierarchical breakdown of all revenue accounts
        5. Expense section with hierarchical breakdown of all expense accounts
        6. Net profit calculation
        7. Professional styling with appropriate use of color, typography, and spacing
        8. Format all monetary values as Indonesian Rupiah (Rp) with thousand separators
        
        Return ONLY the full HTML code that I can save directly as a standalone HTML file.
        """
    else:  # balance-sheet-ifrs
        prompt = f"""
        Create a visually appealing Balance Sheet in an HTML format that I can save as a PDF. 
        The report should look professional and use a clean design with the following data:
        
        {data_json}
        
        The report should include:
        1. Company header with "PT. Jurnal by Mekari" as the company name
        2. Title "Laporan Neraca Keuangan"
        3. Date information showing the as of date
        4. Assets section with hierarchical breakdown of all asset accounts
        5. Liabilities section with hierarchical breakdown of all liability accounts
        6. Equity section with hierarchical breakdown of all equity accounts
        7. Total liabilities and equity calculation
        8. Professional styling with appropriate use of color, typography, and spacing
        9. Format all monetary values as Indonesian Rupiah (Rp) with thousand separators
        
        Return ONLY the full HTML code that I can save directly as a standalone HTML file.
        """
    
    message = client.messages.create(
        model="claude-3-opus-20240229",  
        max_tokens=4000,
        temperature=0.2,  
        system="You are a professional financial report designer. Create beautiful, standards-compliant HTML/CSS reports that look like they were made by a professional designer.",
        messages=[{"role": "user", "content": prompt}]
    )
    
    html_content = message.content[0].text
    
    if "```html" in html_content:
        html_content = html_content.split("```html")[1].split("```")[0].strip()
    elif "```" in html_content:
        html_content = html_content.split("```")[1].split("```")[0].strip()
    
    return html_content

    
    




    